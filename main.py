print("Symulacja ruchu ciał niebieskich z wykorzystaniem Praw Keplera")     
print("Kinga Witschenbach")
print("Wiktoria Czechońska")

#WSTĘP
print("Na wstępie zapoznamy się z tematem, po przez przybliżenie czym są ciała niebieskie, \n kim jest Johannes Kepler i na czy polegają jego Prawa.")
print("\n\n")

#KOD 1
cialaniebieskie = ("Ciało niebieskie \n każdy naturalny obiekt fizyczny oraz układ powiązanych ze sobą obiektów lub ich struktur, występujący w przestrzeni kosmicznej poza granicą atmosfery ziemskiej.\n Ciała niebieskie są przedmiotem zainteresowania astronomii.\n Wśród ciał niebieskich wyróżnia się obiekty, układy i struktury.\n Hipotetyczne ciało niebieskie rozmiarów planety to obiekt synestialny.")
johanneskepler = ("Johannes Kepler – niemiecki matematyk, astronom i astrolog, jedna z czołowych postaci rewolucji naukowej w XVII wieku.\n Najbardziej znany jest z nazwanych jego nazwiskiem praw ruchu planet, skodyfikowanych przez późniejszych astronomów na podstawie jego prac Astronomia nova, Harmonices Mundi i Epitome astronomiae Copernicanae.\n Prawa te wykorzystano do potwierdzenia słuszności teorii grawitacji Isaaca Newtona.")
ciekawostka1 =("Światło dochodzi w 8 minut ze Słońca do Ziemi.\n")
ciekawostka2 =("Na Wenus zamiast opadów śniegu, są opady metalu.")
numery=[]
print("Jeśli chcesz dowiedzieć się co to ciała niebieskie wybierz 1")
print("Jeśli chcesz dowiedzieć się kim był Joannes Kepler wybierz 2")
print("Jeśli chcesz zobczyć ciekawostki wybierz 3")
print("Jeśli chcesz zakończyć należy wpisać 0")

while numery != 0 :
    numery = int(input("\nWybierz o czym chcesz się dowiedzieć : "))

    if (numery == 1 ):
        print (cialaniebieskie)
    elif (numery == 2 ):
        print (johanneskepler)
    elif (numery == 3 ):
        print(ciekawostka1,ciekawostka2)
    elif (numery == 0 ):
        break
    else :
        print("Błąd")

print ("KONIEC\n")

#KOD 2
prawo1 = ("Każda planeta Układu Słonecznego porusza się wokół Słońca po orbicie w kształcie elipsy, w której w jednym z ognisk jest Słońce.")
prawo2 = ("W równych odstępach czasu promień wodzący planety, poprowadzony od Słońca, zakreśla równe pola.")
prawo3 = ("Stosunek kwadratu okresu obiegu planety wokół Słońca do sześcianu wielkiej półosi jej orbity jest stały dla wszystkich planet w Układzie Słonecznym")

numery=[]
print ("Jeśli Chcesz zobaczyć jedno z praw trzeba wpisać numer (1,2,3) jeśli chcesz zakończyć należy wpisać 0 ")
while numery != 0 :
    numery = int(input("\nWybierz które prawo chcesz zobczyć: \n"))

    if (numery == 1 ):
        print (prawo1)
    elif (numery == 2 ):
        print (prawo2)
    elif (numery == 3 ):
        print (prawo3)
    elif (numery == 0):
        break
    else :
        print("Błąd")

print ("KONIEC\n")

#KOD 3


from openpyxl import load_workbook
wb = load_workbook('dane_planety.xlsx')
sheet = wb.active


# Dane- Ziemia
ez = sheet['C1'].value
Tz = sheet['A1'].value
az = sheet['B1'].value

# Dane- Merkury
em = sheet['C2'].value
Tm = sheet['A2'].value
am = sheet['B2'].value

# Dane- Jowisz
ej = sheet['C3'].value
Tj = sheet['A3'].value
aj = sheet['B3'].value

# definicja funkci na odległość między środkiem elipsy a ogniskiem (c=ea) i wybór przez użytkownika, dla której planety chce to policzyc
# + pytanie, czy chce policzyc dla kolejnej planety czy przejść dalej


numery = []
print("Korzystając z I Prawa Keplera możemy obliczyć odległość między środkiem elipsy a jej ogniskiem dla wybranej planety")
print("Jeśli chcesz poznać tę wielkość dla Ziemi wybierz 1, dla Merkurego-2, dla Jowisza- 3 a jeśli chcesz zakończyć wpisz 0 ")
while numery != 0:
    numery = int(input("\nWpisz numer planety, którą wybierasz: \n"))

    if (numery == 1):
        e = ez
        a = az
        float(a)
        float(e)
        c = e * a
        print("Odległość między środkiem elipsy a jej ogniskiem dla Ziemi wynosi: ", c)
    elif (numery == 2):
        e = em
        a = am
        float(a)
        float(e)
        c = e * a
        print("Odległość między środkiem elipsy a jej ogniskiem dla Merkurego wynosi: ", c)

    elif (numery == 3):
        e = ej
        a = aj
        float(a)
        float(e)
        c = e * a
        print("Odległość między środkiem elipsy a jej ogniskiem dla Jowisza wynosi: ", c)

    elif (numery == 0):
        break
    else:
        print("Błąd")

print("KONIEC\n")

#WYKRES 

import matplotlib.pyplot as plt

y = [ez*az,em*am,ej*aj]
x = ['Ziemia', 'Mars', 'Jowisz']
plt.plot(x,y)
plt.plot(x,y,color='r',lw = 2, ls='-')
plt.title('Wykres przedstawiający odległość między środkiem \n elipsy a jej ogniskiem dla danej planety ')
plt.grid(True)
plt.show()

#KOD 4

from openpyxl import load_workbook
wb = load_workbook('dane_planety.xlsx')
sheet = wb.active

# Dane- Ziemia
ez = sheet['C1'].value
Tz = sheet['A1'].value
az = sheet['B1'].value

# Dane- Merkury
em = sheet['C2'].value
Tm = sheet['A2'].value
am = sheet['B2'].value

# Dane- Jowisz
ej = sheet['C3'].value
Tj = sheet['A3'].value
aj = sheet['B3'].value

numery =[]

#KOD 5

print("Korzystając z III Prawa Keplera można po przez przyrównanie dwóch planet obliczyć okres jednej z nich, \n ponieważ stosunek kwadratu okresu planety do sześcianu wielkiej półosi jej orbity jest stały dla wszystkich planet w Układzie Słonecznym\n ")
print("Jeśli chcesz poznać tę wielkość dla Ziemi wybierz 1, dla Merkurego 2, dla Jowisza 3 a jeśli chcesz zakończyć wpisz 0 \n")

import math
while True:
    try:
        numery = int(input("\nWpisz numer planety, którą wybierasz: "))
        
    except ValueError:
        print("Wprowadzono błędny typ danych, spróbuj ponownie")
 
    if (numery == 1):
        T= Tm**2
        a= (az ** 3/ am ** 3)
        float(a)
        float(T)
        Tz = math.sqrt(T*a)
        print("\nOkres dla Ziemi przyrównując ją do Merkurego wynosi: ", Tz)

        T = Tj ** 2
        a = (az ** 3 / aj ** 3)
        float(a)
        float(T)
        Tz = math.sqrt(T*a)
        print("\nOkres dla Ziemi przyrównując ją do Jowisza wynosi: ", Tz)
        print("Z prawa tego wynika, że im większa orbita, tym dłuższy okres obiegu, oraz że prędkość liniowa na orbicie jest odwrotnie proporcjonalna do pierwiastka promienia orbity")

    elif (numery == 2):

        T = Tj ** 2
        a = (am ** 3 / aj ** 3)
        float(a)
        float(T)
        Tm = math.sqrt(T*a)
        print("\nOkres dla Merkurego przyrównując go do Jowisza wynosi: ", Tm)

        T = Tz ** 2
        a = (am ** 3 / az ** 3)
        float(a)
        float(T)
        Tm = math.sqrt(T*a)
        print("\nOkres dla Merkurego przyrównując go do Ziemii wynosi: ", Tm)
        print("Z prawa tego wynika, że im większa orbita, tym dłuższy okres obiegu, oraz że prędkość liniowa na orbicie jest odwrotnie proporcjonalna do pierwiastka promienia orbity")

    elif (numery == 3):

        T = Tm ** 2
        a = (aj ** 3 / am ** 3)
        float(a)
        float(T)
        Tj = math.sqrt(T*a)
        print("\nOkres dla Jowisza przyrównując go do Merkurego wynosi: ", Tj)

        T = Tz ** 2
        a = (aj ** 3 / az ** 3)
        float(a)
        float(T)
        Tj = math.sqrt(T*a)
        print("\n Okres dla Jowisza przyrównując go do Ziemii wynosi: ", Tj)
        print("Z prawa tego wynika, że im większa orbita, tym dłuższy okres obiegu, oraz że prędkość liniowa na orbicie jest odwrotnie proporcjonalna do pierwiastka promienia orbity")

    elif (numery == 0):
        break
    else:
        print("\nBłąd")

print("KONIEC\n")

#WYKRES 2

import matplotlib.pyplot as plt
import math
from pylab import *

Tzm = math.sqrt((Tm**2)*(az ** 3/ am ** 3))
Tzj = math.sqrt((Tj**2)*(az ** 3/ aj ** 3))
Tmj = math.sqrt((Tj**2)*(am ** 3/ aj ** 3))
Tmz = math.sqrt((Tz**2)*(am ** 3/ az ** 3))
Tjm = math.sqrt((Tm**2)*(aj ** 3/ am ** 3))
Tjz = math.sqrt((Tz**2)*(aj ** 3/ az ** 3))

y = [Tzm/Tzj,Tmj/Tmz,Tjm/Tjz]
x = ['Ziemia','Merkury','Jowisz']
plt.plot(x,y)
axis(['Ziemia', 'Jowisz', -5, 5])
plt.plot(x,y,color='g',lw = 1, ls='-')
plt.title('Wykres przedstawiający stosunek okresów obiegu wokół Słońca dla danej planety \n przyrównując ją do dwóch pozostałych planet. ')
plt.grid(True)
plt.show()

#MINI GRA
print("Czas na sprawdzenie wiedzy o Kosmosie!\n")
print("Zasady gry są proste: \nWybierasz poziom trudności gry w którą chcesz zagrać (Prostą czy Trudną)")
print("\nW obu grach należy odpowiedzieć na jedno pytanie.")
print("\nBędzie liczył się za równo czas w którym odpowiecie prawidłowo na pytanie jak i ilość prób.")


grac = []
while grac != "N" :

    print("\n(P=prosta/T=Trudna/N=Nie chcę grać)")
    grac = input("\nCzas wybrać poziom trudności gry: \n")

    if (grac == "P") :
        def czy_a_jest_duzo_mniejsze_od_b(a, b):
            if b - a > 2000:
                return True
            else:
                return False


        def czy_a_jest_mniejsze_od_b(a, b):
            if b - a > 100:
                return True
            else:
                return False


        def czy_a_jest_troche_mniejsze_od_b(a, b):
            if b - a > 10:
                return True
            else:
                return False

        def czy_a_jest_minimalnie_mniejsze_od_b(a, b):
            if b-a > 0:
                return True
            else:
                return False

        def czy_a_jest_minimalnie_wieksze_od_b(a, b):
            if a-b > 0:
                return True
            else:
                return False

        def czy_a_jest_troche_wieksze_od_b(a, b):
            if a - b > 10:
                return True
            else:
                return False


        def czy_a_jest_wieksze_od_b(a, b):
            if a - b > 100:
                return True
            else:
                return False


        def czy_a_jest_duzo_wieksze_od_b(a, b):
            if a - b > 2000:
                return True
            else:
                return False


        import time
        start = time.time()


        def start_gry():
            prawidlowa = 6371
            szukana = -1
            ilosc = 0
            while prawidlowa != szukana:
                print("Ile wynosi promień Ziemii w km: ")
                szukana = int(input())

                if czy_a_jest_duzo_mniejsze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest duzo mniejsza od prawidłowej")

                if czy_a_jest_mniejsze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest mniejsza od prawidłowej")

                if czy_a_jest_troche_mniejsze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest troche mniejsza od prawidłowej")

                if czy_a_jest_minimalnie_mniejsze_od_b(szukana, prawidlowa):
                    print("Zgadnięta jest minimalnie mniejsza od prawidłowej")

                if czy_a_jest_minimalnie_wieksze_od_b(szukana,prawidlowa):
                    print("Zgadnieta jest minimalnie wieksza od prawidłowej")

                if czy_a_jest_troche_wieksze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest troche wieksza od prawidłowej")

                if czy_a_jest_wieksze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest wieksza od prawidłowej")

                if czy_a_jest_duzo_wieksze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest duzo wieksza od prawidłowej")

                ilosc += 1

            print(f"Gratulacje \nIlosc podjetych prob {ilosc}")


        start_gry()

        end = time.time()
        t = end - start
        print(f"Twój czas wynosi {t}")


    elif (grac == "T") :

        def czy_a_jest_o_wiele_mniejsze_od_b(a, b):
            if b - a > 50000:
                return True
            else:
                return False


        def czy_a_jest_duzo_mniejsze_od_b(a, b):
            if b - a > 5000:
                return True
            else:
                return False


        def czy_a_jest_mniejsze_od_b(a, b):
            if b - a > 1000:
                return True
            else:
                return False


        def czy_a_jest_troche_mniejsze_od_b(a, b):
            if b - a > 10:
                return True
            else:
                return False

        def czy_a_jest_minimalnie_mniejsze_od_b(a, b):
            if b-a > 0:
                return True
            else:
                return False

        def czy_a_jest_minimalnie_wieksze_od_b(a, b):
            if a-b > 0:
                return True
            else:
                return False

        def czy_a_jest_troche_wieksze_od_b(a, b):
            if a - b > 10:
                return True
            else:
                return False


        def czy_a_jest_wieksze_od_b(a, b):
            if a - b > 1000:
                return True
            else:
                return False


        def czy_a_jest_duzo_wieksze_od_b(a, b):
            if a - b > 5000:
                return True
            else:
                return False


        def czy_a_jest_o_wiele_wieksze_od_b(a, b):
            if a - b > 50000:
                return True
            else:
                return False


        import time
        start = time.time()

        def start_gry():
            prawidlowa = 69911
            szukana = -1
            ilosc = 0
            while prawidlowa != szukana:
                print("Ile wynosi promień Jowisza w km?")
                szukana = int(input())
                if czy_a_jest_o_wiele_mniejsze_od_b(szukana, prawidlowa):
                    print("Zgadnięta jest o wiele mniejsza od prawidłowej")

                if czy_a_jest_duzo_mniejsze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest duzo mniejsza od prawidłowej")

                if czy_a_jest_mniejsze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest mniejsza od prawidłowej")

                if czy_a_jest_minimalnie_mniejsze_od_b(szukana, prawidlowa):
                    print("Zgadnięta jest minimalnie mniejsza od prawidłowej")

                if czy_a_jest_minimalnie_wieksze_od_b(szukana,prawidlowa):
                    print("Zgadnieta jest minimalnie wieksza od prawidłowej")

                if czy_a_jest_troche_mniejsze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest troche mniejsza od prawidłowej")

                if czy_a_jest_troche_wieksze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest troche wieksza od prawidłowej")

                if czy_a_jest_wieksze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest wieksza od prawidłowej")

                if czy_a_jest_duzo_wieksze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest duzo wieksza od prawidłowej")

                if czy_a_jest_o_wiele_wieksze_od_b(szukana, prawidlowa):
                    print("Zgadnieta jest o wiele wieksza od prawidlowej")

                ilosc += 1

            print(f"Gratulacje \nIlosc podjetych prob {ilosc}")


        start_gry()

        end = time.time()
        t = end - start
        print(f"Twój czas wynosi {t}")

    elif (grac == "N"):
        break

    else :
        print("Wystąpił Błąd")


print("Koniec Gry!\n")

#SYMULACJA LUB WYKRES PRZEDSTAWIAJĄCY RUCH PLANET 
import matplotlib as mpl
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import numpy as np


mpl.rcParams['legend.fontsize'] = 10
fig = plt.figure()
ax = Axes3D(fig)

theta = np.linspace(-4 * np.pi, 4 * np.pi, 100)
z = np.linspace(0, 365.242, 100)
r = 149600000
x = r * np.sin(theta)
y = r * np.cos(theta)
ax.plot(x, y, z, label="Ruch Ziemii wokół Słońca")
ax.legend()

plt.show()
#ciekawe zakończenie programu 
